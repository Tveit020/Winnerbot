from urllib.request import urlopen
import pandas as pd
import time
from datetime import date, timedelta
import matplotlib.pyplot as plt
import numpy as np
from src.scoreclass import Scoreclass
import openpyxl
from openpyxl.styles import PatternFill
from src.export_to_drive import GoogleDriveManager

# URL of the website to scrape

# url = "http://olympus.realpython.org/profiles/dionysus"
# headers = {
#     "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/128.0.0.0 Safari/537.36"
# }


def read_url(url_in):
    page = urlopen(url_in)
    html = page.read().decode("utf-8")
    # dumbass site doesn't give me truth values correctly
    html = html.replace("true", "True")
    html = html.replace("false", "False")
    dict_out = eval(html)

    return dict_out


def export_and_highlight(scoreclass, wrong_picks, excel_path="out/output.xlsx"):
    scoreclass.df.to_excel(excel_path, index=False)

    # Open the Excel file for formatting
    wb = openpyxl.load_workbook(excel_path)
    ws = wb.active

    # Find column indices for "Week" and each player
    header = [cell.value for cell in ws[1]]
    week_col_idx = header.index("Week") + 1  # openpyxl is 1-based

    # Highlight wrong picks in red
    red_fill = PatternFill(
        start_color="FFFF0000", end_color="FFFF0000", fill_type="solid"
    )

    for week_num, failures in wrong_picks.items():
        # Find the row for this week
        for row in range(2, ws.max_row + 1):
            if ws.cell(row=row, column=week_col_idx).value == str(week_num):
                for player in failures:
                    if player in header:
                        col_idx = header.index(player) + 1
                        ws.cell(row=row, column=col_idx).fill = red_fill

    # Highlight "Total" row (Row 20) in red if any value > 3
    total_row = 20
    for col_idx in range(1, ws.max_column + 1):
        cell_value = ws.cell(row=total_row, column=col_idx).value
        # Only check numeric cells (skip header, etc.)
        if isinstance(cell_value, (int, float)) and cell_value >= 3:
            ws.cell(row=total_row, column=col_idx).fill = red_fill

    wb.save(excel_path)
    wb.save("output/output.xlsx")


# Send an HTTP GET request to the website
# response = requests.get(url, headers=headers)
if __name__ == "__main__":

    scoreclass = Scoreclass()
    folder_id = "1GZvN3wq27RqxPQc3di0PrHebwOPPoZrY"
    scoreclass.df = pd.read_csv(scoreclass.df_path)

    total_days = date.today() - date(2025, 9, 3)
    # total_days = date.today() - date(2025, 8, 3)
    iter_num = total_days.days + 1
    # Index through the 7 days of the week(including today)
    # for i in range(iter_num):
    wrong_picks = {}  # Dictionary to track wrong picks for each player
    previous_week = 0
    for i in range(iter_num):

        # game_date = date.today() + timedelta(-50 + i)
        game_date = date(2025, 9, 3) + timedelta(i)
        # game_date = date(2025, 8, 3) + timedelta(i)

        url = scoreclass.base_url + game_date.strftime("%Y%m%d")
        html_dict = read_url(url)

        games = html_dict["events"]
        if games == []:
            continue
        week_num = games[0]["week"]["number"]
        s = scoreclass.df.iloc[week_num - 1]
        print(week_num)

        # Check for missing picks for this week using scoreclass.df
        # Only check the most current week if its Monday has passed
        monday_of_week = game_date - timedelta(days=game_date.weekday()) + timedelta(7)
        if (previous_week != week_num) and (date.today() >= monday_of_week):
            previous_week = week_num
            for player in scoreclass.df.columns:
                if player in ["Week", "Total"]:
                    continue
                pick = s[player]
                if pd.isna(pick) or pick == "":
                    scoreclass.df.loc[len(scoreclass.df) - 1, player] = (
                        int(scoreclass.df.loc[len(scoreclass.df) - 1, player]) + 1
                    )
                    if week_num not in wrong_picks:
                        wrong_picks[week_num] = []
                    wrong_picks[week_num].append(player)
        # Print or use the wrong_picks dictionary as needed

        for game in games:
            for team in game["competitions"][0]["competitors"]:
                team_ind = team["team"]["abbreviation"]
                if "winner" in team:
                    win = team["winner"]
                else:
                    print("Game has not been played yet")
                    continue
                if win == False:
                    failures = s[s == team_ind].index.tolist()
                    print(f"{team_ind} lost")
                    for failure in failures:
                        scoreclass.df.loc[len(scoreclass.df) - 1, failure] = (
                            int(scoreclass.df.loc[len(scoreclass.df) - 1, failure]) + 1
                        )
                        # Track wrong picks
                        if week_num not in wrong_picks:
                            wrong_picks[week_num] = []
                        wrong_picks[week_num].append(failure)

        # Check for missing picks for this week using scoreclass.df

        print("Wrong picks this round:", wrong_picks)

        time.sleep(0.25)
    print(scoreclass.df)
    print(wrong_picks)
    scoreclass.df.to_csv(f"output/Scores Week {week_num}.csv", index=False)
    # Output DataFrame to Excel
    export_and_highlight(
        scoreclass, wrong_picks, excel_path=f"output/Scores Week {week_num}.xlsx"
    )
    gdrive = GoogleDriveManager()
    gdrive.upload_file("output/output.xlsx", folder_id=folder_id)
